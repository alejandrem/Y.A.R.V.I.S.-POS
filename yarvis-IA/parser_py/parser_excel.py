"""
Parser de catálogos en formato Excel (.xlsx, .xls).
Detecta automáticamente columnas por nombre o posición.
"""
import io
from fastapi import HTTPException
from core.utils import limpiar_producto, es_categoria

try:
    import openpyxl
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False


def parsear_excel(bytes_excel: bytes) -> list[dict]:
    """
    Parsea un catálogo en formato Excel (.xlsx, .xls).
    
    Detecta automáticamente las columnas:
    - Busca columnas con "nombre", "producto", "descripción"
    - Busca columnas con "costo", "precio"
    - Busca columnas con "venta", "publico"
    - Busca columnas con "categoría", "tipo"
    """
    if not OPENPYXL_DISPONIBLE:
        raise HTTPException(status_code=500, detail="openpyxl no está instalado. Ejecuta: pip install openpyxl")
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(bytes_excel), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer Excel: {str(e)}")
    
    productos = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Leer todas las filas
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        
        # Detectar headers y sus índices
        headers = [str(h).lower().strip() if h else '' for h in rows[0]]
        
        # Verificar si la primera fila parece header
        es_header = False
        for h in headers:
            if h and any(k in h for k in ['nombre', 'producto', 'name', 'cost', 'price', 'venta', 'costo', 'precio', 'categor', 'category', 'tipo', 'stock']):
                es_header = True
                break
        
        # Si la primera fila parece ser datos, intentar buscar headers en la segunda fila
        data_start = 0 if not es_header else 1
        if not es_header and len(rows) > 1:
            second_headers = [str(v).lower().strip() if v else '' for v in rows[1]]
            for h in second_headers:
                if h and any(k in h for k in ['nombre', 'producto', 'name', 'cost', 'price', 'venta', 'costo', 'precio', 'categor', 'category', 'tipo', 'stock']):
                    headers = second_headers
                    data_start = 2
                    break
        
        # Buscar columnas por nombre
        col_nombre = None
        col_costo = None
        col_venta = None
        col_categoria = None
        col_stock = None
        
        for i, h in enumerate(headers):
            if not h:
                continue
            if any(k in h for k in ['nombre', 'producto', 'descripción', 'descripcion', 'articulo', 'artículo', 'name', 'product', 'description', 'item']):
                col_nombre = i
            elif any(k in h for k in ['costo', 'cost', 'precio_compra', 'precio costo', 'cost price']):
                col_costo = i
            elif any(k in h for k in ['venta', 'publico', 'público', 'precio_venta', 'precio venta', 'precio', 'price', 'selling', 'sale', 'retail']):
                col_venta = i
            elif any(k in h for k in ['categoría', 'categoria', 'tipo', 'seccion', 'sección', 'departamento', 'category', 'type', 'section']):
                col_categoria = i
            elif any(k in h for k in ['stock', 'existencia', 'cantidad', 'inventario', 'quantity', 'inventory']):
                col_stock = i
        
        # Si no encontramos nombre, detectar por posición
        if col_nombre is None:
            for i, h in enumerate(headers):
                if h and not any(c.isdigit() for c in h.replace('$', '').replace(',', '')):
                    col_nombre = i
                    break
        
        # Detectar categoría si no se encontró por nombre
        if col_categoria is None:
            for i, h in enumerate(headers):
                if h and any(k in h for k in ['categoría', 'categoria', 'tipo', 'seccion', 'sección', 'departamento', 'category', 'type', 'section']):
                    col_categoria = i
                    break
        
        # Detectar categoría por posición si no se encontró por nombre
        if col_categoria is None and col_nombre is not None:
            for i, h in enumerate(headers):
                if i != col_nombre and h and not any(c.isdigit() for c in h.replace('$', '').replace(',', '')):
                    # Podría ser categoría
                    if len(h) < 30:
                        col_categoria = i
                        break
        
        # Si aún no hay, buscar en las filas de datos (solo si no encontramos nombre)
        if col_nombre is None:
            for row_idx in range(0 if data_start == 1 else data_start, len(rows)):
                row = rows[row_idx]
                if not row:
                    continue
                if col_nombre is None:
                    for i, v in enumerate(row):
                        if v and not isinstance(v, (int, float)):
                            col_nombre = i
                            break
                if col_nombre is not None:
                    break
        
        data_rows = rows[data_start:]
        
        if col_nombre is None:
            continue
        
        # Parsear cada fila
        for row in data_rows:
            if not row or len(row) <= col_nombre:
                continue
            
            nombre = row[col_nombre] if col_nombre < len(row) else None
            if not nombre or not str(nombre).strip():
                continue
            
            nombre_str = str(nombre).strip()
            if es_categoria(nombre_str):
                continue
            
            # Obtener precios (opcionales)
            precio_venta = 0.0
            precio_costo = 0.0
            
            if col_venta is not None:
                try:
                    precio_venta_raw = float(str(row[col_venta]).replace('$', '').replace(',', '').strip()) if col_venta < len(row) and row[col_venta] else 0
                except (ValueError, TypeError):
                    precio_venta_raw = 0
                
                try:
                    precio_costo_raw = float(str(row[col_costo]).replace('$', '').replace(',', '').strip()) if col_costo is not None and col_costo < len(row) and row[col_costo] else 0
                except (ValueError, TypeError):
                    precio_costo_raw = 0
                
                # Detectar orden correcto
                if precio_venta_raw > 0 and precio_costo_raw > 0:
                    if precio_venta_raw < precio_costo_raw:
                        precio_venta = precio_costo_raw
                        precio_costo = precio_venta_raw
                    else:
                        precio_venta = precio_venta_raw
                        precio_costo = precio_costo_raw
                else:
                    precio_venta = precio_venta_raw
                    precio_costo = precio_costo_raw
            
            # Obtener categoría
            categoria = "SIN CATEGORÍA"
            if col_categoria is not None and col_categoria < len(row) and row[col_categoria]:
                categoria = str(row[col_categoria]).strip().upper()
            
            # Obtener stock
            stock = 0
            if col_stock is not None and col_stock < len(row) and row[col_stock]:
                try:
                    stock = int(float(str(row[col_stock]).replace(',', '').strip()))
                except (ValueError, TypeError):
                    stock = 0
            
            # Agregar producto (aunque no tenga precio)
            productos.append({
                "nombre": limpiar_producto(nombre_str),
                "precio_costo": round(precio_costo, 2),
                "precio_venta": round(precio_venta, 2),
                "stock": stock,
                "categoria": categoria,
            })
    
    wb.close()
    return productos
